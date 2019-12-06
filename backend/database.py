from .resources import read_csv, is_empty, trim, trim2, write_csv
from backend.levenshtein import probar_input, compare_by_lenght
from os import path as ospath, getcwd, listdir
import openpyxl


def process_devir_xlsx(ruta):
    tabla = []
    document = openpyxl.load_workbook(ruta)
    for n_hoja in document.sheetnames:
        ws = document[n_hoja]
        for row in ws.iter_rows():
            if type(row[2].value) is int:
                tabla.append({
                    'codigo': row[2].value,
                    'nombre': str(row[3].value).upper(),
                    'precio': row[5].value,
                    'otro': row[1].value,
                })

    document.close()
    return tabla


def process_sd_dist_xlsx(ruta):
    tabla = []
    document = openpyxl.load_workbook(ruta)
    for n_hoja in document.sheetnames:
        ws = document[n_hoja]
        for row in ws.iter_rows():
            if type(row[3].value) is int:
                isbn = '-'
                if type(row[3].value) is int:
                    isbn = str(row[4].value)
                elif row[3].value is not None:
                    isbn = ''.join(row[4].value.split('-'))
                tabla.append({
                    'codigo': row[1].value,
                    'nombre': str(row[2].value).strip().upper(),
                    'precio': float(row[3].value),
                    'ISBN': isbn,
                    'EAN': row[7].value.strip(),
                    'ADENDUM': row[8].value.strip(),
                    'editorial': row[9].value.strip(),
                    'autor': row[10].value.strip()
                })
    document.close()
    return tabla


def process_ivrea_xlsx(ruta):
    tabla = []
    document = openpyxl.load_workbook(ruta)
    for n_hoja in [i for i in document.sheetnames if ('MANGA' in i) or ('COMICS' in i)]:
        ws = document[n_hoja]
        for r in ws.iter_rows():
            if r[2].value is not None and type(r[6].value) is int:
                isbn = '-'
                if type(r[3].value) is int:
                    isbn = str(r[3].value)
                elif r[3].value is not None:
                    isbn = ''.join(r[3].value.split('-'))
                tabla.append({
                    'codigo': r[1].value,
                    'nombre': str(r[2].value).lstrip().rstrip(),
                    'ISBN': isbn,
                    'EAN': r[4].value,
                    'ADENDUM': r[5].value,
                    'precio': float(r[6].value),
                    'agotado': 1 if r[8].value is not None else 0
                })

    document.close()
    return tabla


def process_plan_xlsx(ruta):
    tabla = []
    document = openpyxl.load_workbook(ruta)
    for n_hoja in document.sheetnames:
        rows = list(document[n_hoja].rows)
        for r in rows[9:]:
            if r[0].value is not None:
                t = r[3].value.upper()
                n = t.split(' - ')[0] if 'SIN STOCK' in t else t
                n.strip()

                isbn = '-'
                if type(r[2].value) is int:
                    isbn = str(r[2].value)
                elif r[2].value is not None:
                    isbn = ''.join(r[2].value.split('-'))

                tabla.append({
                    'codigo': r[0].value,
                    'nombre': n,
                    'ISBN': isbn,
                    'precio': float(r[6].value),
                    'agotado': 1 if 'SIN STOCK' in t else 0
                })

    document.close()
    return tabla


def process_mavis(ruta):
    tabla = []
    document = openpyxl.load_workbook(ruta)
    for n_hoja in document.sheetnames:
        ws = document[n_hoja]
        for row in ws.iter_rows():
            if type(row[1].value) is int:
                tabla.append({
                    'nombre': row[0].value.strip().upper(),
                    'precio': float(row[1].value),
                    'otro': row[3].value
                })

    document.close()
    return tabla


def process_pannini(ruta):
    tabla = []
    document = openpyxl.load_workbook(ruta)
    for n_hoja in document.sheetnames:
        ws = document[n_hoja]
        for row in ws.iter_rows():
            if type(row[6].value) is int:
                isbn = '-'
                if type(row[3].value) is int:
                    isbn = str(row[6].value)
                elif row[3].value is not None:
                    isbn = ''.join(row[6].value.split('-'))
                tabla.append({
                    'codigo': row[1].value,
                    'nombre': str(row[2].value).strip().upper(),
                    'precio': float(row[3].value),
                    'ISBN': isbn,
                    'EAN': row[7].value.strip(),
                    'editorial': row[8].value.strip(),
                    'autor': row[9].value.strip()
                })

    document.close()
    return tabla


def open_table(ruta):
    tabla = []
    datos = read_csv(ruta)
    header = datos[0]
    rows = datos[1:]
    for i, row in enumerate(rows):
        tabla.append({})
        for j, value in enumerate(row):
            head = header[j]
            if head == 'codigo':
                try:
                    value = int(value)
                except ValueError:
                    # sd_dist codes are not numeric.
                    pass
            elif head in ('precio', 'precio_siva', 'descuento'):
                value = float(value)
            tabla[i].update({header[j]: value})
    return tabla


root = getcwd() + '/data'
tablas = []
for fd in listdir(root):
    route = ospath.join(root, fd)
    for file in listdir(route):
        path = ospath.join(route, file)
        if fd == 'DEVIR':
            tablas.append(process_devir_xlsx(path))
        elif fd == 'SD':
            tablas.append(process_sd_dist_xlsx(path))
        elif fd == 'IVREA':
            tablas.append(process_ivrea_xlsx(path))
        elif fd == 'PLAN T':
            tablas.append(process_plan_xlsx(path))
        elif fd == 'PANNINI':
            tablas.append(process_pannini(path))
        elif fd == 'MAVIS':
            tablas.append(process_mavis(path))


def select_many(key, columns):
    results = []
    for tabla in tablas:
        for row in tabla:
            if key in row['nombre']:
                d = {'nombre': row['nombre']}
                for column in columns:
                    if column in row:
                        d[column] = row[column]
                results.append(d)

    if not len(results):
        cercana = {}
        for tabla in tablas:
            names = [row['nombre'] for row in tabla]
            cercana.update(probar_input(key, names))

        for c in compare_by_lenght(key, cercana.pop(min(cercana))):
            results.extend(select_many(c, columns))

    results.sort(key=lambda o: o['nombre'])
    return results
